"""
app/multimodal/documents.py — Extract text from PDF, DOCX, XLSX attachments.

Supported: .pdf (pypdf), .docx (python-docx), .xlsx/.xls (openpyxl).
Returns None for unsupported types or on error.
"""

import asyncio
import tempfile
from typing import Callable, Optional

from loguru import logger
from pyrogram import Client
from pyrogram.types import Message

MAX_CHARS = 8000


async def extract_document_text(client: Client, message: Message) -> Optional[str]:
    """Download document and extract plain text. Returns None if unsupported."""
    doc = message.document
    if not doc:
        return None

    fname = (doc.file_name or "").lower()
    extractor: Optional[Callable[[str], str]] = None

    if fname.endswith(".pdf"):
        extractor = _pdf
    elif fname.endswith(".docx"):
        extractor = _docx
    elif fname.endswith((".xlsx", ".xls")):
        extractor = _xlsx

    if extractor is None:
        logger.debug(f"[DOCS] Unsupported file type: {fname!r}")
        return None

    try:
        with tempfile.TemporaryDirectory() as tmp:
            path = await client.download_media(message, file_name=f"{tmp}/{doc.file_name or 'doc'}")
            if not path:
                return None
            loop = asyncio.get_event_loop()
            text = await loop.run_in_executor(None, extractor, path)
            if text:
                logger.info(f"[DOCS] Extracted {len(text)} chars from {fname!r}")
            return text[:MAX_CHARS] if text else None

    except Exception as e:
        logger.error(f"[DOCS] Extraction error ({fname!r}): {e}")
        return None


def _pdf(path: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    return "\n".join(page.extract_text() or "" for page in reader.pages).strip()


def _docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            line = "\t".join(str(c) for c in row if c is not None)
            if line.strip():
                rows.append(line)
    return "\n".join(rows)
